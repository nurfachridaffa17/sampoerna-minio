from minio import Minio
import os
import openpyxl
from dotenv.main import load_dotenv
import boto3
from .log import Logger


# Create a client with the MinIO server playground, its access key

load_dotenv()

log = Logger()

client = Minio(
    os.environ['URL_MINIO'],
    access_key=os.environ['ACCESS_KEY_MINIO'],
    secret_key=os.environ['SECRET_KEY_MINIO'],
    secure=False  # Set to False if using an insecure connection (not recommended for production)
)

s3_client = boto3.client(
    's3',
    endpoint_url=os.environ['URL_S3'],
    aws_access_key_id=os.environ['S3_ACCESS_KEY'],
    aws_secret_access_key=os.environ['S3_SECRET_KEY'],
    region_name=os.environ['S3_REGION']
)

def download_file(bucket_name, object_name, local_directory):
    try:
        # Check if the bucket exists
        if client.bucket_exists(bucket_name):
            # Ensure the local_directory exists, create if it doesn't
            os.makedirs(local_directory, exist_ok=True)

            # Generate the file path for the downloaded file
            filename = object_name.split("/")[-1]
            local_file_path = os.path.join(local_directory, filename)

            # Download the file from MinIO
            client.fget_object(bucket_name, object_name, local_file_path)

            log.info(f"File downloaded successfully to {local_file_path}")
        else:
            log.error(f"Bucket '{bucket_name}' does not exist")
    except Exception as e:
        log.error(f"Error downloading file: {str(e)}")

def file_exists_in_local(file_path):
    return os.path.exists(file_path)


# def upload_file_to_s3(bucket_name, file_path, object_name):
#     try:
#         s3_client.upload_file(file_path, bucket_name, object_name)
#         log.info(f"File uploaded successfully to {bucket_name}/{object_name}")
#     except Exception as e:
#         log.error(f"Error uploading file: {str(e)}")

def check_object_existence(bucket_name, object_name):
    try:
        response = s3_client.head_object(Bucket=bucket_name, Key=object_name)
        return True
    except s3_client.exceptions.NoSuchKey:
        return False
    except Exception as e:
        print(f"Error checking object existence: {str(e)}")
        return False


def upload_file_to_s3(bucket_name, file_path, object_name):
    try:
        if not check_object_existence(bucket_name, object_name):
            s3_client.upload_file(file_path, bucket_name, object_name)
            log.info(f"File uploaded successfully to {bucket_name}/{object_name}")
        else:
            log.info(f"File already exists in {bucket_name}/{object_name}")
    except Exception as e:
        log.error(f"Error uploading file: {str(e)}")



path = os.environ['PATH_FOLDER_EXCEL']
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row 
m_column = sheet_obj.max_column

for i in range(1, m_row + 1):
    path_url = sheet_obj.cell(row = i, column = 1).value
    try:
        bucket_name_minio = os.environ['BUCKET_NAME_MINIO']
        object_name_minio = path_url
        local_file_path = "/opt/sampoerna-minio/static"

        minio_file_path = os.path.join(local_file_path, os.path.basename(path_url))

        if not file_exists_in_local(minio_file_path):
            download_file(bucket_name_minio, object_name_minio, local_file_path)
            log.info("Download file from MinIO success", path_url)
        else:
            log.info("File already exists in local static folder, skipping download", path_url)
    except Exception as e:
        log.error(f"Error downloading file: {str(e)}")
    
    try:
        bucket_name_s3 = os.environ['S3_BUCKET']
        object_name_s3 = path_url
        file_path = "/opt/sampoerna-minio/static/{}".format(path_url.split("/")[-1])

        if not check_object_existence(bucket_name_s3, object_name_s3):
            upload_file_to_s3(bucket_name_s3, file_path, object_name_s3)
            log.info("Upload file to S3 success", path_url)
        else:
            log.info("Object already exists in S3, skipping upload", path_url)
    except Exception as e:
        log.error(f"Error uploading file: {str(e)}")
    
    try:
        os.remove(file_path)
        log.info("Delete file success", path_url)
    except Exception as e:
        log.error(f"Error deleting file: {str(e)}")

log.info("All process success")